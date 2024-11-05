from fastapi import FastAPI, File, UploadFile, Form
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
from io import BytesIO
import openpyxl
import uvicorn
import holidays
import base64
from datetime import datetime

app = FastAPI()

# Configurar CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Lista de rutas
RUTAS = [
    'C6-001', 'C6-002', 'C6-002A', 'C6-003', 'C6-004', 'C6-005', 'C6-005A', 'C6-006C',
    'C6-007', 'C6-008', 'C6-009', 'C6-010C', 'C6-010B', 'C6-011', 'C6-012', 'C6-012A', 'C6-013',
    'C6-014', 'C6-015A', 'C6-016', 'C6-016A', 'C6-018', 'C6-019', 'C6-019A', 'C6-020',
    'C6-021', 'C6-022', 'C6-022A', 'C6-023', 'C6-024', 'C6-025', 'C6-025B'
]

def obtener_factor_dia(fecha, country_holidays):
    if fecha.weekday() < 5:  # Lunes a Viernes
        if fecha.date() not in country_holidays:
            return 209  # Día hábil
        return 106  # Festivo (si cae en día laboral)
    elif fecha.weekday() == 5:  # Sábado
        return 173
    else:  # Domingo
        return 106

def obtener_tipo_dia(fecha, country_holidays):
    if fecha.date() in country_holidays:
        return 'festivo'
    elif fecha.weekday() == 5:
        return 'sabado'
    elif fecha.weekday() == 6:
        return 'domingo'
    else:
        return 'habil'

def convert_to_datetime(df, column_name):
    """Convierte una columna específica a datetime de manera segura."""
    try:
        df = df.copy()
        df = df[~df[column_name].isin(['Prom. Pax x Mes', 'Prom. Km x Mes'])]
        df[column_name] = pd.to_datetime(df[column_name], errors='coerce')
        mask_nat = df[column_name].isna()
        if mask_nat.any():
            df.loc[mask_nat, column_name] = pd.to_datetime(
                df.loc[mask_nat, column_name],
                format='%d-%m-%y',
                errors='coerce'
            )
        return df
    except Exception as e:
        print(f"Error en la conversión de fecha: {e}")
        return df

def process_km_sheet(km_sheet, single_date, datos_fecha):
    km_sheet['A1'] = single_date.strftime('%Y-%m-%d')
    for j, ruta in enumerate(RUTAS, start=4):
        km = datos_fecha[datos_fecha['Ruta'] == ruta]['Km'].sum()
        km_sheet[f'B{j}'] = km
    return km_sheet

def process_usuarios_sheet(usuarios_sheet, single_date, datos_fecha):
    usuarios_sheet['A1'] = single_date.strftime('%Y-%m-%d')
    for j, ruta in enumerate(RUTAS, start=4):
        usuarios = datos_fecha[datos_fecha['Ruta'] == ruta]['Pax'].sum()
        usuarios_sheet[f'B{j}'] = usuarios
    return usuarios_sheet

@app.post("/upload/")
async def upload_excel(
    file: UploadFile = File(...),
    start_date: str = Form(...),
    end_date: str = Form(...)
):
    try:
        contents = await file.read()
        excel_data = pd.ExcelFile(BytesIO(contents))

        df_usuarios = pd.read_excel(excel_data, sheet_name='Data Usuarios')
        df_kilometros = pd.read_excel(excel_data, sheet_name='Data Kilometros')

        if 'Fecha' not in df_usuarios.columns:
            df_usuarios = df_usuarios.rename(columns={df_usuarios.columns[1]: 'Fecha'})
        if 'Pax' not in df_usuarios.columns:
            df_usuarios = df_usuarios.rename(columns={df_usuarios.columns[7]: 'Pax'})
        if 'Ruta' not in df_usuarios.columns:
            df_usuarios = df_usuarios.rename(columns={df_usuarios.columns[6]: 'Ruta'})

        if 'Fecha' not in df_kilometros.columns:
            df_kilometros = df_kilometros.rename(columns={df_kilometros.columns[1]: 'Fecha'})
        if 'Ruta' not in df_kilometros.columns:
            df_kilometros = df_kilometros.rename(columns={df_kilometros.columns[6]: 'Ruta'})
        if 'Km' not in df_kilometros.columns:
            df_kilometros = df_kilometros.rename(columns={df_kilometros.columns[7]: 'Km'})

        df_usuarios = convert_to_datetime(df_usuarios, 'Fecha')
        df_kilometros = convert_to_datetime(df_kilometros, 'Fecha')

        start_date_dt = pd.to_datetime(start_date)
        end_date_dt = pd.to_datetime(end_date)
        date_range = pd.date_range(start=start_date_dt, end=end_date_dt)

        plantilla = openpyxl.load_workbook('PlantillaSegunda.xlsx')

        if len(date_range) > 7:
            return JSONResponse(
                content={"error": "El rango de fechas no puede exceder 7 días."},
                status_code=400
            )

        excel_data_dict = {}
        country_holidays = holidays.CountryHoliday('CO')

        for i, single_date in enumerate(date_range):
            tipo_dia = obtener_tipo_dia(single_date, country_holidays)
            
            # Procesar hoja de kilómetros
            if tipo_dia == 'habil':
                km_sheet_name = f"KILOMETROSHABIL{i + 1}"
            elif tipo_dia == 'sabado':
                km_sheet_name = "KILOMETROS6"
            elif tipo_dia == 'domingo':
                km_sheet_name = "KILOMETROS7"
            else:
                km_sheet_name = "KILOMETROS8"

            if km_sheet_name in plantilla.sheetnames:
                km_sheet = plantilla[km_sheet_name]
                datos_fecha_km = df_kilometros[df_kilometros['Fecha'].dt.date == single_date.date()]
                km_sheet = process_km_sheet(km_sheet, single_date, datos_fecha_km)
                excel_data_dict[km_sheet_name] = [
                    [cell.value for cell in row] for row in km_sheet.iter_rows()
                ]

            # Procesar hoja de usuarios
            if tipo_dia == 'habil':
                usuarios_sheet_name = f"USUARIOSHABIL{i + 1}"
            elif tipo_dia == 'sabado':
                usuarios_sheet_name = "USUARIOS6"
            elif tipo_dia == 'domingo':
                usuarios_sheet_name = "USUARIOS7"
            else:
                usuarios_sheet_name = "USUARIOS8"

            if usuarios_sheet_name in plantilla.sheetnames:
                usuarios_sheet = plantilla[usuarios_sheet_name]
                datos_fecha_usuarios = df_usuarios[df_usuarios['Fecha'].dt.date == single_date.date()]
                usuarios_sheet = process_usuarios_sheet(usuarios_sheet, single_date, datos_fecha_usuarios)
                excel_data_dict[usuarios_sheet_name] = [
                    [cell.value for cell in row] for row in usuarios_sheet.iter_rows()
                ]

            # Procesar hoja original
            sheet_name = f"Hoja{i + 1}"
            if sheet_name in plantilla.sheetnames:
                sheet = plantilla[sheet_name]
                sheet['A1'] = single_date.strftime('%Y-%m-%d')

                filtered_usuarios = df_usuarios[df_usuarios['Fecha'].dt.date == single_date.date()]
                filtered_kilometros = df_kilometros[df_kilometros['Fecha'].dt.date == single_date.date()]

                if not filtered_usuarios.empty:
                    total_usuarios = filtered_usuarios['Pax'].sum() / 2
                    sheet['B8'] = total_usuarios

                if not filtered_kilometros.empty:
                    total_kilometros = filtered_kilometros['Km'].sum()
                    sheet['B3'] = total_kilometros

                    factor_dia = obtener_factor_dia(single_date, country_holidays)
                    sheet['A3'] = total_kilometros / factor_dia

                    sheet['B5'] = total_kilometros - (total_kilometros * 0.02)
                    sheet['A5'] = sheet['B5'].value / factor_dia

                excel_data_dict[sheet_name] = [
                    [cell.value for cell in row] for row in sheet.iter_rows()
                ]

        for sheet_name in plantilla.sheetnames:
            if sheet_name not in excel_data_dict:
                sheet = plantilla[sheet_name]
                sheet.sheet_state = 'hidden'

        buffer = BytesIO()
        plantilla.save(buffer)
        buffer.seek(0)
        excel_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')

        return JSONResponse(content={
            "excel_data": excel_data_dict,
            "excel_file": excel_base64,
            "message": "Archivo procesado con éxito"
        })

    except Exception as e:
        print(f"Error detallado: {str(e)}")
        return JSONResponse(
            content={"error": f"Error procesando archivo: {str(e)}"},
            status_code=500
        )

if __name__ == "__main__":
    print("Servidor iniciado")
    uvicorn.run(app, host="0.0.0.0", port=8001)